"""
logger_serial_ke_excel.py

TUJUAN:
Membaca data JSON yang dikirim ESP32-S3 lewat Serial (USB), lalu
LANGSUNG menuliskannya ke file Excel (.xlsx) baris demi baris,
real-time, selama alat menyala -- tanpa perlu copy-paste manual
dari Serial Monitor.

CARA PAKAI:
1. Tutup Arduino IDE Serial Monitor dulu (satu port cuma bisa dipakai
   satu program dalam satu waktu -- kalau Serial Monitor Arduino IDE
   masih terbuka, script ini akan gagal connect).
2. Cek nama port ESP32 kamu di Device Manager (Windows), biasanya
   berbentuk "COM3", "COM5", dst. Update variabel SERIAL_PORT di bawah.
3. Jalankan: python logger_serial_ke_excel.py
4. Biarkan berjalan selama alat menyala. Setiap baris JSON yang masuk
   otomatis ditambahkan sebagai baris baru di file Excel.
5. Tekan CTRL+C untuk berhenti. File Excel aman tersimpan.

CATATAN PENTING:
- Selama script ini berjalan, JANGAN buka file Excel-nya di Excel
  secara bersamaan di Windows -- Excel mengunci file saat dibuka,
  script akan gagal menulis. Tutup dulu file Excel-nya, baru jalankan
  script, atau buka file itu SETELAH kamu berhenti logging.
- Kalau ingin melihat data yang sedang masuk tanpa menutup logger,
  buka file CSV pendampingnya (dibuat otomatis, nama sama, akhiran
  .csv) -- CSV tidak dikunci Excel selama tidak dibuka juga.
"""

import serial
import json
import csv
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ======================= KONFIGURASI =======================
SERIAL_PORT = "COM5"      # GANTI sesuai port ESP32 kamu (lihat Device Manager)
BAUD_RATE = 115200         # Harus sama dengan Serial.begin() di main.ino
OUTPUT_XLSX = "data_vibris.xlsx"
OUTPUT_CSV = "data_vibris.csv"   # cadangan, aman dibuka saat logger jalan
# =============================================================

# Urutan kolom mengikuti field JSON yang dikirim RaspberryPiDataTransmitter.cpp
KOLOM = [
    "waktu_lokal", "rms_v", "rms_x", "rms_y", "rms_z", "rms_a",
    "cur", "cur_raw_adc", "temp", "temp_raw", "rpm", "severity",
    "status", "e_unbalance", "e_misalign", "e_bpfo", "e_bpfi",
    "diagnosis", "diag_conf"
]


def siapkan_file_excel(path):
    """Buat file Excel baru dengan header kalau belum ada, atau lanjutkan
    dari file yang sudah ada supaya tidak menimpa data lama."""
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetri"
    ws.append(KOLOM)
    wb.save(path)
    return wb, ws


def siapkan_file_csv(path):
    file_baru = not os.path.exists(path)
    f = open(path, "a", newline="", encoding="utf-8")
    writer = csv.writer(f)
    if file_baru:
        writer.writerow(KOLOM)
    return f, writer


def main():
    print(f"[LOGGER] Menyambungkan ke {SERIAL_PORT} @ {BAUD_RATE} baud...")
    ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=2)
    print("[LOGGER] Tersambung. Menunggu data JSON dari ESP32...")

    wb, ws = siapkan_file_excel(OUTPUT_XLSX)
    csv_file, csv_writer = siapkan_file_csv(OUTPUT_CSV)

    baris_masuk = 0
    try:
        while True:
            raw_line = ser.readline().decode("utf-8", errors="ignore").strip()

            # Baris yang bukan JSON (misal baris TELEMETRI MONITORING atau
            # baris [FFT] debug) dilewati -- kita cuma ambil baris JSON.
            if not raw_line.startswith("{"):
                continue

            try:
                data = json.loads(raw_line)
            except json.JSONDecodeError:
                # Baris JSON kepotong / rusak di tengah transmisi -- lewati
                # saja, jangan sampai bikin script crash.
                continue

            waktu = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            baris = [
                waktu,
                data.get("rms_v"), data.get("rms_x"), data.get("rms_y"), data.get("rms_z"),
                data.get("rms_a"), data.get("cur"), data.get("cur_raw_adc"),
                data.get("temp"), data.get("temp_raw"), data.get("rpm"),
                data.get("severity"), data.get("status"),
                data.get("e_unbalance"), data.get("e_misalign"),
                data.get("e_bpfo"), data.get("e_bpfi"),
                data.get("diagnosis"), data.get("diag_conf"),
            ]

            # Tulis ke Excel
            ws.append(baris)
            wb.save(OUTPUT_XLSX)   # simpan tiap baris masuk -- aman kalau listrik/alat mati mendadak

            # Tulis ke CSV juga (cadangan, bisa dipantau live tanpa lock Excel)
            csv_writer.writerow(baris)
            csv_file.flush()

            baris_masuk += 1
            print(f"[LOGGER] Baris #{baris_masuk} tersimpan | status={data.get('status')} rpm={data.get('rpm')}")

    except KeyboardInterrupt:
        print("\n[LOGGER] Dihentikan oleh user. Menyimpan file terakhir...")
    except serial.SerialException as e:
        print(f"[LOGGER] ERROR koneksi serial: {e}")
        print("[LOGGER] Cek: port benar? Arduino Serial Monitor sudah ditutup? Kabel USB tersambung?")
    finally:
        wb.save(OUTPUT_XLSX)
        csv_file.close()
        ser.close()
        print(f"[LOGGER] Selesai. Total {baris_masuk} baris tersimpan di {OUTPUT_XLSX} dan {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
