import sys
import os
import csv
import json
import time
from datetime import datetime
from collections import deque

# Import Serial & Threading untuk Integrasi Komunikasi Data Hardware ESP32
try:
    import serial
    import serial.tools.list_ports
    import threading
except ImportError:
    serial = None

# Import openpyxl untuk fitur "Export ke Excel" (opsional; kalau belum
# terinstall, tombol export akan menampilkan pesan cara install-nya alih-alih
# bikin program crash)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QHBoxLayout, QVBoxLayout,
    QGridLayout, QStackedWidget, QFrame, QListWidget, QComboBox, QMessageBox,
    QSlider, QDialog
)
from PyQt5.QtCore import QTimer, Qt
from PyQt5.QtGui import QFont
import pyqtgraph as pg

# ===================== KONFIGURASI OPERASIONAL =====================
# Menyesuaikan dengan port USB-Enhanced-SERIAL CH343 laptop kamu yang terdeteksi.
# Kalau port ini salah/berubah (mis. pindah laptop, ganti kabel USB, atau
# dipindah ke Raspberry Pi), dashboard akan otomatis MENCARI SENDIRI port
# yang paling mirip ESP32 (lihat fungsi _resolve_serial_port di bawah),
# jadi tidak perlu selalu diedit manual.
SERIAL_PORT = 'COM6'
BAUD_RATE = 115200
LOG_DIR = "logs"
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)

# Kata kunci deskripsi USB-Serial chip yang umum dipakai board ESP32
# (CH340/CH343, CP210x, FTDI). Dipakai untuk auto-detect port kalau
# SERIAL_PORT di atas tidak ditemukan / salah.
ESP32_USB_HINTS = ["CH340", "CH343", "CP210", "USB-SERIAL", "USB SERIAL", "FTDI", "SILICON LABS"]

# ===================== PALET WARNA INDUSTRI VIBRIS =====================
COL_BG_MAIN = "#2d3135"      
COL_PANEL_DARK = "#1c1e22"   
COL_ACCENT = "#2a6f97"       
COL_TEXT_LIGHT = "#f2f2f2"
COL_OK = "#2e7d32"
COL_WARN = "#e08e00"
COL_BAD = "#c62828"
COL_HEADER_BG = "#e3b419"

# Ambang batas D^2 Mahalanobis PERSIS SAMA dengan firmware ESP32
# (lihat MahalanobisDetector.cpp: CHI_SQUARE_95 / CHI_SQUARE_99), supaya
# garis ambang batas di grafik dashboard konsisten dengan status yang
# dikirim device, bukan angka karang-karangan dari sisi Python.
D2_THRESHOLD_WASPADA = 9.49   # chi-square, df=4, confidence 95%
D2_THRESHOLD_BAHAYA = 13.28   # chi-square, df=4, confidence 99%

# Urutan tingkat keparahan status, dipakai buat menentukan "Kondisi Terparah" per sesi
STATUS_SEVERITY = {"normal": 0, "waspada": 1, "bahaya": 2}

class Dashboard(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("HMI | Rotating Machinery Detection System")
        
        # ===== PENGATURAN UKURAN TETAP 480x320 SESUAI LAYAR TFT RASPBERRY PI =====
        # Dikunci pas (bukan fullscreen otomatis) supaya tampilan selalu presisi
        # 480x320, gak tergantung resolusi layar yang lagi dipakai buat testing
        # di laptop (yang biasanya lebih gede dari layar TFT Raspi aslinya).
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setFixedSize(480, 320)
        self.setStyleSheet(f"background-color: {COL_BG_MAIN}; color: {COL_TEXT_LIGHT}; font-family: Arial;")

        # Data Buffer Sinkronisasi Plot Grafik (Panjang data 50 point antrean)
        self.data_len = 50
        self.time_buffer = deque(maxlen=self.data_len)
        self.v_buffer = deque(maxlen=self.data_len)
        self.vx_buffer = deque(maxlen=self.data_len)  # Getaran sumbu X
        self.vy_buffer = deque(maxlen=self.data_len)  # Getaran sumbu Y
        self.vz_buffer = deque(maxlen=self.data_len)  # Getaran sumbu Z
        self.a_buffer = deque(maxlen=self.data_len)
        self.cur_buffer = deque(maxlen=self.data_len)
        self.temp_buffer = deque(maxlen=self.data_len)
        self.rpm_buffer = deque(maxlen=self.data_len)
        self.d2_buffer = deque(maxlen=self.data_len)
        self.tick = 0
        self.last_processed_tick = 0
        self.last_raw_line = ""

        # State playback halaman detail rekaman (Logs & Saves -> buka rekaman,
        # ditampilkan INLINE di dashboard yang sama, bukan window terpisah)
        self.logdet_times = []
        self.logdet_v_vals = []
        self.logdet_a_vals = []
        self.logdet_cur_vals = []
        self.logdet_t_vals = []
        self.logdet_play_index = 0
        self.logdet_play_speed = 1.0
        self.logdet_timer = QTimer(self)
        self.logdet_timer.timeout.connect(self._logdet_step)

        # Variabel Penampung Nilai Sensor Terkini (Awalnya kosong sebelum ada data serial masuk)
        self.current_v = None
        self.current_vx = None
        self.current_vy = None
        self.current_vz = None
        self.current_a = None
        self.current_cur = None
        self.current_temp = None
        self.current_rpm = None
        self.current_d2 = None
        self.current_status_device = ""

        # ===== STATISTIK SESI (dipakai tab Processed & Summary, di-reset via tombol RESET) =====
        self.session_sample_count = 0
        self.session_rpm_sum = 0.0
        self.session_d2_max = 0.0
        self.session_worst_status = "Normal"
        self.session_waspada_count = 0
        self.session_bahaya_count = 0
        self.anomaly_events = []  # list of str, event log kejadian Waspada/Bahaya

        # Status Operasional Recording & Serial Hardware
        self.recording = False
        self.csv_file = None
        self.csv_writer = None
        self.ser = None
        self.serial_connected = False

        # Root Layout Utama Vertikal
        root = QVBoxLayout(self)
        root.setContentsMargins(4, 4, 4, 4)
        root.setSpacing(4)

        # ===== HEADER BAR =====
        header = QHBoxLayout()
        header.setSpacing(6)
        header_frame = QFrame()
        header_frame.setStyleSheet(f"background-color: {COL_HEADER_BG}; border-radius: 4px;")
        header_frame.setLayout(header)

        self.machine_combo = QComboBox()
        self.machine_combo.addItems([
            "- Pilih Mesin Target -",
            "Blower Industri UMKM",
            "Motor Induksi Pompa Air",
            "Kompresor Production",
            "Mesin Blender"
        ])
        self.machine_combo.setFixedHeight(24)
        self.machine_combo.setStyleSheet(
            "background-color: rgba(0,0,0,0.15); color: #ffffff; font-size: 11px; "
            "font-weight: bold; padding: 2px 4px; border: 1px solid #1c1e22; border-radius: 3px;"
        )
        self.machine_combo.currentTextChanged.connect(self._on_machine_changed)
        header.addWidget(self.machine_combo, 3)

        header.addStretch(1)

        self.btn_reset = QPushButton("RESET")
        self.btn_reset.setStyleSheet(f"background-color: {COL_ACCENT}; color: #ffffff; font-weight: bold; font-size: 8px; padding: 4px 8px; border-radius: 3px;")
        self.btn_reset.clicked.connect(self._reset_session)
        header.addWidget(self.btn_reset)

        self.btn_debug = QPushButton("DEBUG")
        self.btn_debug.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: #ffffff; font-weight: bold; font-size: 8px; padding: 4px 8px; border-radius: 3px;")
        self.btn_debug.clicked.connect(self._show_debug_info)
        header.addWidget(self.btn_debug)

        self.lbl_conn_dot = QLabel("●")
        self.lbl_conn_dot.setStyleSheet(f"font-size: 12px; font-weight: bold; color: {COL_BAD};")
        header.addWidget(self.lbl_conn_dot)

        self.time_lbl = QLabel("--:--:--")
        self.time_lbl.setStyleSheet("font-size: 10px; font-weight: bold; color: #1c1e22;")
        self.time_lbl.setAlignment(Qt.AlignRight)
        header.addWidget(self.time_lbl)

        root.addWidget(header_frame)

        # ===== STACKED PAGES (KONTEN UTAMA DI TENGAH) DENGAN NAVIGASI PANAH KIRI/KANAN =====
        stack_row = QHBoxLayout()
        stack_row.setSpacing(2)

        self.btn_nav_prev = QPushButton("‹")
        self.btn_nav_prev.setFixedWidth(28)
        self.btn_nav_prev.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: #ffffff; font-size: 16px; font-weight: bold; border-radius: 14px;")
        self.btn_nav_prev.clicked.connect(lambda: self._change_page((self.stack.currentIndex() - 1) % 4))
        stack_row.addWidget(self.btn_nav_prev)

        self.stack = QStackedWidget()
        self.stack.addWidget(self._page_raw())       # Index 0
        self.stack.addWidget(self._page_recording()) # Index 1
        self.stack.addWidget(self._page_processed()) # Index 2
        self.stack.addWidget(self._page_summary())   # Index 3
        self.stack.addWidget(self._page_log_detail()) # Index 4 (dibuka dari Logs & Saves, inline, bukan window baru)
        stack_row.addWidget(self.stack, 1)

        self.btn_nav_next = QPushButton("›")
        self.btn_nav_next.setFixedWidth(28)
        self.btn_nav_next.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: #ffffff; font-size: 16px; font-weight: bold; border-radius: 14px;")
        self.btn_nav_next.clicked.connect(lambda: self._change_page((self.stack.currentIndex() + 1) % 4))
        stack_row.addWidget(self.btn_nav_next)

        root.addLayout(stack_row, 1)

        # ===== NAVIGASI TOMBOL UTAMA DI SEBELAH BAWAH (WARNA HURUF HITAM CONTRAS) =====
        nav_bottom = QHBoxLayout()
        nav_bottom.setSpacing(4)
        
        self.btn_raw = QPushButton("RAW READING")
        self.btn_rec = QPushButton("LOGS SAVES")
        self.btn_proc = QPushButton("PROCESSED")
        self.btn_sum = QPushButton("SUMMARY")

        self.menu_buttons = [self.btn_raw, self.btn_rec, self.btn_proc, self.btn_sum]
        for i, btn in enumerate(self.menu_buttons):
            btn.setFixedHeight(42)  # Tinggi ergonomis ramah sentuhan jari pada layar kecil
            btn.setStyleSheet(self._menu_style(False))
            btn.clicked.connect(lambda checked, idx=i: self._change_page(idx))
            nav_bottom.addWidget(btn)
        
        root.addLayout(nav_bottom)

        # ===== MENGAKTIFKAN LOGIKA UTARA SENSOR SERIAL =====
        self._init_serial_connection()

        # ===== TIMER UTAMA REFRESH GUI (Per 200 milidetik) =====
        self.main_timer = QTimer(self)
        self.main_timer.timeout.connect(self._update_gui)
        self.main_timer.start(200) 

        # Set Halaman Awal & Muat Berkas Rekaman (.csv) bawaan ketua tim
        self._change_page(0)
        self._refresh_log_list()
        self.show()

    def _menu_style(self, active):
        # Seluruh tulisan teks pada tombol menu dipastikan tetap berwarna hitam pekat kontras tinggi
        if active:
            return f"background-color: {COL_ACCENT}; color: #000000; font-size: 12px; font-weight: bold; border: 1px solid white; border-radius: 4px;"
        return f"background-color: #cfcfcf; color: #000000; font-size: 12px; font-weight: bold; border: 1px solid #444; border-radius: 4px;"

    def _change_page(self, idx):
        self.stack.setCurrentIndex(idx)
        # Halaman detail log (index 4) dianggap masih "bagian" dari tab
        # Logs & Saves (index 1) buat keperluan highlight tombol menu bawah,
        # karena cuma bisa dibuka dari sana.
        highlight_idx = 1 if idx == 4 else idx
        for i, btn in enumerate(self.menu_buttons):
            btn.setStyleSheet(self._menu_style(i == highlight_idx))

    # ===================== HALAMAN 1: RAW READING (4 KOTAK GRAFIK INDIVIDU) =====================
    def _page_raw(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(1, 1, 1, 1)
        layout.setSpacing(1)

        # Catatan: dropdown "Pilih Mesin Target" sudah dipindah ke bar kuning
        # (header) di atas, jadi baris terpisah di sini dihapus supaya ruang
        # vertikal halaman ini sepenuhnya dipakai buat grafik.

        main_raw = QHBoxLayout()
        main_raw.setSpacing(2)
        
        # Grid layout untuk membagi porsi layar menjadi 4 bagian grafik sub-plot dedicated
        layout_grafik_grid = QGridLayout()
        layout_grafik_grid.setSpacing(2)
        layout_grafik_grid.setContentsMargins(0, 0, 0, 0)
        
        pg.setConfigOptions(antialias=True)

        def _tidy_plot(plot_widget, title):
            """Rapikan tampilan grafik: font sumbu lebih besar & jarak antar
            angka sumbu diperlebar supaya gak numpuk/bertabrakan di layar
            kecil 480x320. Dipakai sama untuk keempat grafik biar konsisten."""
            axis_font = QFont("Arial", 8)
            plot_item = plot_widget.getPlotItem()
            for axis_name in ("left", "bottom"):
                axis = plot_item.getAxis(axis_name)
                axis.setStyle(tickFont=axis_font, tickTextOffset=6, autoExpandTextSpace=True)
                axis.setTextPen(pg.mkPen('#dddddd'))
            # Batasi jumlah angka yang muncul di sumbu X (waktu) supaya gak
            # numpuk berdempetan — dipperlebar lagi jadi 1 angka tiap 15 tick
            # (dari sebelumnya tiap 10), karena grafik makin besar tapi
            # jumlah titik data (50) tetap sama.
            plot_item.getAxis("bottom").setTickSpacing(major=15, minor=15)
            plot_item.getAxis("left").setWidth(40)
            plot_widget.setTitle(title, size="10pt")

        # 1. Grafik Getaran (Vibration)
        self.graph_v = pg.PlotWidget(title="Vibration (G)")
        self.graph_v.setBackground(COL_PANEL_DARK)
        self.graph_v.showGrid(x=True, y=True, alpha=0.2)
        self.curve_v = self.graph_v.plot(pen=pg.mkPen('#ff4d4d', width=2))
        _tidy_plot(self.graph_v, "Vibration (G)")
        layout_grafik_grid.addWidget(self.graph_v, 0, 0)
        
        # 2. Grafik Suara (Sound)
        self.graph_a = pg.PlotWidget(title="Sound (dB)")
        self.graph_a.setBackground(COL_PANEL_DARK)
        self.graph_a.showGrid(x=True, y=True, alpha=0.2)
        self.curve_a = self.graph_a.plot(pen=pg.mkPen('#4da6ff', width=2))
        _tidy_plot(self.graph_a, "Sound (dB)")
        layout_grafik_grid.addWidget(self.graph_a, 0, 1)
        
        # 3. Grafik Arus Listrik (Current)
        self.graph_cur = pg.PlotWidget(title="Current (A)")
        self.graph_cur.setBackground(COL_PANEL_DARK)
        self.graph_cur.showGrid(x=True, y=True, alpha=0.2)
        self.curve_cur = self.graph_cur.plot(pen=pg.mkPen('#ffeb3b', width=2))
        _tidy_plot(self.graph_cur, "Current (A)")
        layout_grafik_grid.addWidget(self.graph_cur, 1, 0)
        
        # 4. Grafik Suhu (Temperature)
        self.graph_temp = pg.PlotWidget(title="Temp (°C)")
        self.graph_temp.setBackground(COL_PANEL_DARK)
        self.graph_temp.showGrid(x=True, y=True, alpha=0.2)
        self.curve_temp = self.graph_temp.plot(pen=pg.mkPen('#e040fb', width=2))
        _tidy_plot(self.graph_temp, "Temp (°C)")
        layout_grafik_grid.addWidget(self.graph_temp, 1, 1)
        
        main_raw.addLayout(layout_grafik_grid, 15)

        # Panel Kanan Bagian Nilai Numerik Polosan
        panel_kanan = QVBoxLayout()
        panel_kanan.setSpacing(2)
        
        frame_status = QFrame()
        frame_status.setStyleSheet(f"background-color: {COL_PANEL_DARK}; border-radius: 4px;")
        fs_lay = QVBoxLayout(frame_status)
        fs_lay.setContentsMargins(4, 4, 4, 4)
        
        self.lbl_sys_status = QLabel("● STANDBY")
        self.lbl_sys_status.setStyleSheet(f"font-size: 9px; font-weight: bold; color: #888888;")
        fs_lay.addWidget(self.lbl_sys_status)
        
        self.lbl_val_v = QLabel("Vib: -- G")
        self.lbl_val_vxyz = QLabel("  (X: -- | Y: -- | Z: -- G)")  # nilai mentah sebelum diakarkan jadi Vib
        self.lbl_val_a = QLabel("Snd: -- dB")
        self.lbl_val_cur = QLabel("Cur: -- A")
        self.lbl_val_temp = QLabel("Tmp: -- °C")
        self.lbl_val_rpm = QLabel("RPM: --")
        self.lbl_val_d2 = QLabel("D² (Severity): --")  # dipersingkat dari "D²(Mahalanobis)" biar gak kepotong

        for lbl in [self.lbl_val_v, self.lbl_val_vxyz, self.lbl_val_a, self.lbl_val_cur, self.lbl_val_temp,
                    self.lbl_val_rpm, self.lbl_val_d2]:
            lbl.setStyleSheet("font-size: 11px; color: #dddddd; padding: 2px 0px;")
            lbl.setWordWrap(True)
            fs_lay.addWidget(lbl)
        fs_lay.setSpacing(4)

        panel_kanan.addWidget(frame_status)
        main_raw.addLayout(panel_kanan, 2)

        layout.addLayout(main_raw)
        return page

    # ===================== HALAMAN 2: LOGS & SAVES (KONTROL REKAM DATA) =====================
    def _page_recording(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(4)

        rec_ctrl = QHBoxLayout()
        self.btn_toggle_rec = QPushButton("MULAI RECORDING")
        self.btn_toggle_rec.setStyleSheet("background-color: #cfcfcf; color: #000000; font-weight: bold; font-size: 9px; height: 24px; border-radius: 4px;")
        self.btn_toggle_rec.clicked.connect(self._toggle_recording)
        
        self.lbl_rec_status = QLabel("● Device Standby")
        self.lbl_rec_status.setStyleSheet("font-size: 8px; color: #aaa;")
        
        rec_ctrl.addWidget(self.btn_toggle_rec, 2)
        rec_ctrl.addWidget(self.lbl_rec_status, 1)
        layout.addLayout(rec_ctrl)

        lbl_daftar = QLabel("Daftar Rekaman Mesin (.CSV):")
        lbl_daftar.setStyleSheet("font-size: 8px;")
        layout.addWidget(lbl_daftar)
        
        self.log_list = QListWidget()
        self.log_list.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: white; font-size: 9px;")
        # Klik langsung pada file rekaman -> otomatis membuka panel baru berisi grafik hasil deteksi
        self.log_list.itemDoubleClicked.connect(self._open_log_detail)
        layout.addWidget(self.log_list)

        btn_lay = QHBoxLayout()
        self.btn_watch_log = QPushButton("Buka Panel Deteksi")
        self.btn_export_excel = QPushButton("Export ke Excel")
        self.btn_delete_log = QPushButton("Hapus Rekaman")

        self.btn_watch_log.setStyleSheet("background-color: #cfcfcf; color: #000000; font-size: 9px; height: 20px; font-weight: bold; border-radius: 3px;")
        self.btn_export_excel.setStyleSheet("background-color: #217346; color: #ffffff; font-size: 9px; height: 20px; font-weight: bold; border-radius: 3px;")
        self.btn_delete_log.setStyleSheet("background-color: #cfcfcf; color: #000000; font-size: 9px; height: 20px; font-weight: bold; border-radius: 3px;")

        self.btn_watch_log.clicked.connect(self._open_log_detail_from_button)
        self.btn_export_excel.clicked.connect(self._export_selected_log_to_excel)
        self.btn_delete_log.clicked.connect(self._delete_selected_log)

        btn_lay.addWidget(self.btn_watch_log)
        btn_lay.addWidget(self.btn_export_excel)
        btn_lay.addWidget(self.btn_delete_log)
        layout.addLayout(btn_lay)

        lbl_hint = QLabel("* Klik salah satu file rekaman di atas untuk membuka panel hasil deteksi\n"
                           "  (grafik Vibration/Sound/Current/Temp beserta status diagnosanya).")
        lbl_hint.setStyleSheet("font-size: 8px; color: #888; font-style: italic; margin-top: 4px;")
        lbl_hint.setWordWrap(True)
        layout.addWidget(lbl_hint)

        return page

    # ===================== HALAMAN 3: PROCESSED READING (RPM, MAHALANOBIS D², LOG ANOMALI) =====================
    def _page_processed(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(4)

        self.lbl_proc_snapshot = QLabel("Menunggu data... | RPM: -- | D²: --")
        self.lbl_proc_snapshot.setStyleSheet("font-size: 8px; color: #999;")
        layout.addWidget(self.lbl_proc_snapshot)

        # ===== 2 GRAFIK: RPM ESTIMASI & MAHALANOBIS D² (dengan garis ambang batas) =====
        grid = QGridLayout()
        grid.setSpacing(4)
        pg.setConfigOptions(antialias=True)

        self.graph_rpm = pg.PlotWidget(title="RPM Estimasi")
        self.graph_rpm.setBackground(COL_PANEL_DARK)
        self.graph_rpm.showGrid(x=True, y=True, alpha=0.3)
        self.curve_rpm = self.graph_rpm.plot(pen=pg.mkPen('#4da6ff', width=1.5))
        grid.addWidget(self.graph_rpm, 0, 0)

        self.graph_d2 = pg.PlotWidget(title="Mahalanobis D²")
        self.graph_d2.setBackground(COL_PANEL_DARK)
        self.graph_d2.showGrid(x=True, y=True, alpha=0.3)
        self.curve_d2 = self.graph_d2.plot(pen=pg.mkPen('#ff6666', width=1.5))
        # Garis putus-putus ambang batas Waspada (kuning) & Bahaya (merah),
        # nilainya persis sama dengan chi-square threshold di firmware.
        line_waspada = pg.InfiniteLine(pos=D2_THRESHOLD_WASPADA, angle=0,
                                        pen=pg.mkPen(COL_WARN, width=1.5, style=Qt.DashLine))
        line_bahaya = pg.InfiniteLine(pos=D2_THRESHOLD_BAHAYA, angle=0,
                                       pen=pg.mkPen(COL_BAD, width=1.5, style=Qt.DashLine))
        self.graph_d2.addItem(line_waspada)
        self.graph_d2.addItem(line_bahaya)
        grid.addWidget(self.graph_d2, 0, 1)

        layout.addLayout(grid, 3)

        # ===== LOG KEJADIAN ANOMALI SEPANJANG SESI =====
        lbl_anomali_title = QLabel("Log Kejadian Anomali:")
        lbl_anomali_title.setStyleSheet("font-size: 8px; font-weight: bold; color: #ccc;")
        layout.addWidget(lbl_anomali_title)

        self.list_anomali = QListWidget()
        self.list_anomali.setStyleSheet("background-color: #ffffff; color: #222222; font-size: 9px;")
        self.list_anomali.addItem("Tidak ada kejadian anomali sepanjang sesi ini.")
        layout.addWidget(self.list_anomali, 2)

        # ===== RINGKASAN STATISTIK SESI (footer hijau muda) =====
        self.lbl_session_summary = QLabel(
            "Sesi: 0 sample | RPM rata-rata: 0.0 | D² max: 0.00 | Kondisi terparah: Normal | Waspada: 0x, Bahaya: 0x."
        )
        self.lbl_session_summary.setStyleSheet(
            "font-size: 9px; color: #1c3d1c; background-color: #d7f0d7; padding: 4px; border-radius: 3px;"
        )
        self.lbl_session_summary.setWordWrap(True)
        layout.addWidget(self.lbl_session_summary)

        return page

    # ===================== HALAMAN 4: DIAGNOSIS SUMMARY (KESIMPULAN KONDISI) =====================
    def _page_summary(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)
        layout.addStretch(1)

        row = QHBoxLayout()
        row.addStretch(1)

        self.box_diagnosis = QFrame()
        self.box_diagnosis.setFixedWidth(360)
        self.box_diagnosis.setStyleSheet(f"background-color: transparent; border: 3px solid #888; border-radius: 4px;")
        box_lay = QVBoxLayout(self.box_diagnosis)
        box_lay.setContentsMargins(14, 14, 14, 14)
        box_lay.setSpacing(8)

        self.lbl_diag_status = QLabel("STATUS: STANDBY")
        self.lbl_diag_status.setStyleSheet("font-size: 20px; font-weight: bold; color: #222;")
        self.lbl_diag_status.setAlignment(Qt.AlignCenter)
        box_lay.addWidget(self.lbl_diag_status)

        self.lbl_diag_readout = QFrame()
        self.lbl_diag_readout.setStyleSheet("border: 2px solid #888; border-radius: 3px;")
        readout_lay = QHBoxLayout(self.lbl_diag_readout)
        readout_lay.setContentsMargins(6, 4, 6, 4)
        self.lbl_diag_rpm = QLabel("RPM: 0.0")
        self.lbl_diag_rpm.setStyleSheet("font-size: 13px; font-weight: bold; color: #222;")
        self.lbl_diag_d2 = QLabel("D²: 0.00")
        self.lbl_diag_d2.setStyleSheet("font-size: 13px; font-weight: bold; color: #222;")
        readout_lay.addWidget(self.lbl_diag_rpm)
        readout_lay.addStretch(1)
        readout_lay.addWidget(self.lbl_diag_d2)
        box_lay.addWidget(self.lbl_diag_readout)

        row.addWidget(self.box_diagnosis)
        row.addStretch(1)
        layout.addLayout(row)
        layout.addStretch(1)

        # Deskripsi ringkas kondisi (dipertahankan, ditaruh di bawah panel besar)
        self.lbl_diag_desc = QLabel("Belum ada data deteksi yang diproses. Silakan pilih target mesin dan hubungkan kabel hardware.")
        self.lbl_diag_desc.setStyleSheet("font-size: 9px; color: #888;")
        self.lbl_diag_desc.setWordWrap(True)
        self.lbl_diag_desc.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.lbl_diag_desc)

        return page

    # ===================== HALAMAN 5 (TERSEMBUNYI): DETAIL REKAMAN (INLINE, BUKAN WINDOW BARU) =====================
    def _page_log_detail(self):
        """Halaman ini cuma bisa diakses lewat 'Logs & Saves' -> klik salah satu
        file rekaman. Ditampilkan di dalam stack yang sama (bukan QDialog/window
        terpisah), jadi playback rekaman langsung muncul di dashboard ini juga."""
        page = QWidget()
        root = QVBoxLayout(page)
        root.setContentsMargins(4, 4, 4, 4)
        root.setSpacing(4)

        self.lbl_logdet_title = QLabel("HASIL DETEKSI REKAMAN: -")
        self.lbl_logdet_title.setStyleSheet("font-size: 10px; font-weight: bold; color: #ffffff;")
        self.lbl_logdet_title.setWordWrap(True)
        root.addWidget(self.lbl_logdet_title)

        # ===== STATUS RINGKAS: LINGKARAN WARNA + NILAI PUNCAK (GANTI KOTAK BESAR) =====
        status_row = QHBoxLayout()
        status_row.setSpacing(6)

        self.lbl_logdet_dot = QLabel("●")
        self.lbl_logdet_dot.setStyleSheet("font-size: 16px; color: #888888;")
        status_row.addWidget(self.lbl_logdet_dot)

        self.lbl_logdet_peak = QLabel("Nilai puncak: -")
        self.lbl_logdet_peak.setWordWrap(True)
        self.lbl_logdet_peak.setStyleSheet("font-size: 9px; color: #cccccc;")
        status_row.addWidget(self.lbl_logdet_peak, 1)

        root.addLayout(status_row)
        root.addSpacing(4)  # sedikit jarak sebelum grafik, biar gak nempel status di atasnya

        # ===== GRID 4 GRAFIK (GAYA SAMA DENGAN TAB RAW READING) =====
        grid = QGridLayout()
        grid.setSpacing(3)
        pg.setConfigOptions(antialias=True)

        self.graph_logdet_v = pg.PlotWidget(title="Vibration (G)")
        self.graph_logdet_v.setBackground(COL_PANEL_DARK)
        self.graph_logdet_v.showGrid(x=True, y=True, alpha=0.2)
        self.curve_logdet_v = self.graph_logdet_v.plot(pen=pg.mkPen('#ff4d4d', width=1.5))
        grid.addWidget(self.graph_logdet_v, 0, 0)

        self.graph_logdet_a = pg.PlotWidget(title="Sound (dB)")
        self.graph_logdet_a.setBackground(COL_PANEL_DARK)
        self.graph_logdet_a.showGrid(x=True, y=True, alpha=0.2)
        self.curve_logdet_a = self.graph_logdet_a.plot(pen=pg.mkPen('#4da6ff', width=1.5))
        grid.addWidget(self.graph_logdet_a, 0, 1)

        self.graph_logdet_cur = pg.PlotWidget(title="Current (A)")
        self.graph_logdet_cur.setBackground(COL_PANEL_DARK)
        self.graph_logdet_cur.showGrid(x=True, y=True, alpha=0.2)
        self.curve_logdet_cur = self.graph_logdet_cur.plot(pen=pg.mkPen('#ffeb3b', width=1.5))
        grid.addWidget(self.graph_logdet_cur, 1, 0)

        self.graph_logdet_temp = pg.PlotWidget(title="Temp (°C)")
        self.graph_logdet_temp.setBackground(COL_PANEL_DARK)
        self.graph_logdet_temp.showGrid(x=True, y=True, alpha=0.2)
        self.curve_logdet_temp = self.graph_logdet_temp.plot(pen=pg.mkPen('#e040fb', width=1.5))
        grid.addWidget(self.graph_logdet_temp, 1, 1)

        # Stretch besar (20) supaya grid grafik memakai HAMPIR SELURUH ruang
        # vertikal yang tersisa, jauh lebih besar dibanding sebelumnya waktu
        # masih berbagi ruang sama kotak status besar.
        root.addLayout(grid, 20)

        # ===== KONTROL PLAYBACK (PLAY/PAUSE, SPEED, SLIDER, KEMBALI) =====
        ctrl = QHBoxLayout()
        ctrl.setSpacing(3)

        self.btn_logdet_play_pause = QPushButton("▶ PLAY")
        self.speed_logdet_combo = QComboBox()
        self.speed_logdet_combo.addItems(["0.5x", "1x", "2x", "4x"])
        self.speed_logdet_combo.setCurrentIndex(1)
        self.btn_logdet_back = QPushButton("◄ KEMBALI")

        self.btn_logdet_play_pause.setStyleSheet("background-color: #cfcfcf; color: #000000; font-weight: bold; font-size: 9px; height: 28px; border-radius: 4px;")
        self.btn_logdet_back.setStyleSheet(f"background-color: {COL_BAD}; color: #ffffff; font-weight: bold; font-size: 9px; height: 28px; border-radius: 4px;")
        self.speed_logdet_combo.setStyleSheet(f"background-color: {COL_PANEL_DARK}; color: white; font-size: 9px; padding: 2px;")

        self.btn_logdet_play_pause.clicked.connect(self._logdet_toggle_play)
        self.speed_logdet_combo.currentIndexChanged.connect(self._logdet_change_speed)
        self.btn_logdet_back.clicked.connect(self._logdet_back)

        self.slider_logdet = QSlider(Qt.Horizontal)
        self.slider_logdet.setMinimum(0)
        self.slider_logdet.setMaximum(0)
        self.slider_logdet.sliderMoved.connect(self._logdet_seek)

        self.lbl_logdet_pos = QLabel("0 / 0")
        self.lbl_logdet_pos.setStyleSheet("font-size: 8px; color: #aaa;")

        ctrl.addWidget(self.btn_logdet_play_pause, 2)
        ctrl.addWidget(self.speed_logdet_combo, 1)
        ctrl.addWidget(self.slider_logdet, 5)
        ctrl.addWidget(self.lbl_logdet_pos, 1)
        ctrl.addWidget(self.btn_logdet_back, 2)
        root.addLayout(ctrl)

        return page

    def _logdet_back(self):
        """Hentikan animasi playback dan kembali ke daftar Logs & Saves."""
        self.logdet_timer.stop()
        self._change_page(1)

    def _logdet_load_csv(self, filepath):
        """Membaca seluruh isi berkas CSV rekaman menjadi list per kolom sensor."""
        times, v_vals, a_vals, cur_vals, t_vals = [], [], [], [], []
        try:
            with open(filepath, 'r') as f:
                reader = csv.reader(f)
                next(reader)  # lewati baris header
                for i, row in enumerate(reader):
                    times.append(i)
                    v_vals.append(float(row[2]))
                    a_vals.append(float(row[3]))
                    cur_vals.append(float(row[4]))
                    t_vals.append(float(row[5]))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal membaca berkas rekaman: {e}")
        return times, v_vals, a_vals, cur_vals, t_vals

    def _logdet_render_full_overview(self):
        """Menampilkan seluruh isi rekaman sekaligus di grafik (tampilan hasil deteksi penuh)."""
        self.curve_logdet_v.setData(self.logdet_times, self.logdet_v_vals)
        self.curve_logdet_a.setData(self.logdet_times, self.logdet_a_vals)
        self.curve_logdet_cur.setData(self.logdet_times, self.logdet_cur_vals)
        self.curve_logdet_temp.setData(self.logdet_times, self.logdet_t_vals)
        if self.logdet_times:
            self.slider_logdet.blockSignals(True)
            self.slider_logdet.setValue(len(self.logdet_times) - 1)
            self.slider_logdet.blockSignals(False)

    def _logdet_render_diagnosis_summary(self):
        """Menghitung nilai puncak (peak) selama sesi rekaman, lalu menentukan
        warna lingkaran indikator (hijau/jingga/merah) dan teks nilai puncak."""
        if not self.logdet_v_vals:
            self.lbl_logdet_dot.setStyleSheet("font-size: 16px; color: #888888;")
            self.lbl_logdet_peak.setText("Data kosong / tidak bisa dianalisis.")
            return

        peak_v = max(self.logdet_v_vals)
        peak_a = max(self.logdet_a_vals)
        peak_cur = max(self.logdet_cur_vals)
        peak_temp = max(self.logdet_t_vals)

        # Menggunakan ambang batas yang sama seperti diagnosa live pada dashboard utama
        if peak_v > 0.25 or peak_temp > 50.0:
            status, col = "BAHAYA", COL_BAD
        elif peak_v > 0.18 or peak_temp > 42.0:
            status, col = "WASPADA", COL_WARN
        else:
            status, col = "NORMAL", COL_OK

        self.lbl_logdet_dot.setStyleSheet(f"font-size: 16px; color: {col};")
        self.lbl_logdet_peak.setText(
            f"{status} — Puncak: Vib {peak_v:.2f} G | Snd {peak_a:.1f} dB | "
            f"Cur {peak_cur:.2f} A | Tmp {peak_temp:.1f} °C"
        )

    def _logdet_render_frame(self, idx):
        """Menggambar 1 frame animasi (jendela 50 titik terakhir) untuk efek 'muter ulang'."""
        start = max(0, idx - 49)
        t = self.logdet_times[start:idx + 1]
        v = self.logdet_v_vals[start:idx + 1]
        a = self.logdet_a_vals[start:idx + 1]
        c = self.logdet_cur_vals[start:idx + 1]
        tp = self.logdet_t_vals[start:idx + 1]

        self.curve_logdet_v.setData(t, v)
        self.curve_logdet_a.setData(t, a)
        self.curve_logdet_cur.setData(t, c)
        self.curve_logdet_temp.setData(t, tp)

        self.slider_logdet.blockSignals(True)
        self.slider_logdet.setValue(idx)
        self.slider_logdet.blockSignals(False)
        self.lbl_logdet_pos.setText(f"{idx + 1} / {len(self.logdet_times)}")

    def _logdet_step(self):
        if self.logdet_play_index >= len(self.logdet_times):
            self.logdet_timer.stop()
            self.btn_logdet_play_pause.setText("▶ PLAY")
            return
        self._logdet_render_frame(self.logdet_play_index)
        self.logdet_play_index += 1

    def _logdet_toggle_play(self):
        if not self.logdet_times:
            return
        if self.logdet_timer.isActive():
            self.logdet_timer.stop()
            self.btn_logdet_play_pause.setText("▶ PLAY")
        else:
            # Jika sudah di akhir data, otomatis ulang dari awal
            if self.logdet_play_index >= len(self.logdet_times):
                self.logdet_play_index = 0
            interval_ms = max(20, int(200 / self.logdet_play_speed))
            self.logdet_timer.start(interval_ms)
            self.btn_logdet_play_pause.setText("⏸ PAUSE")

    def _logdet_change_speed(self, combo_idx):
        mapping = {0: 0.5, 1: 1.0, 2: 2.0, 3: 4.0}
        self.logdet_play_speed = mapping.get(combo_idx, 1.0)
        if self.logdet_timer.isActive():
            self.logdet_timer.start(max(20, int(200 / self.logdet_play_speed)))

    def _logdet_seek(self, value):
        self.logdet_play_index = value
        self._logdet_render_frame(value)
    def _init_serial_connection(self):
        if serial is not None:
            # Membuka thread background baru khusus membaca UART secara non-blocking
            t = threading.Thread(target=self._read_serial_worker, daemon=True)
            t.start()
        else:
            print("Library pyserial tidak terinstal atau tidak terdeteksi pada Python interpreter.")

    def _resolve_serial_port(self):
        """Cari port serial yang tepat untuk ESP32.
        Kalau tidak ada port yang cocok, kembalikan None supaya dashboard
        tidak terus mencoba membuka port yang tidak ada."""
        try:
            ports = list(serial.tools.list_ports.comports())
        except Exception:
            ports = []

        if not ports:
            return None

        available = [p.device for p in ports]
        if SERIAL_PORT in available:
            return SERIAL_PORT

        for p in ports:
            desc = f"{p.description} {p.manufacturer or ''}".upper()
            if any(hint in desc for hint in ESP32_USB_HINTS):
                return p.device

        return None

    def _read_serial_worker(self):
        while True:
            try:
                if self.ser is None or not self.ser.is_open:
                    port_to_use = self._resolve_serial_port()
                    if not port_to_use:
                        self.serial_connected = False
                        time.sleep(2)
                        continue

                    self.ser = serial.Serial(port_to_use, BAUD_RATE, timeout=1)
                    self.serial_connected = True

                raw = self.ser.readline()
                if not raw:
                    # Timeout tanpa data baru -> coba lagi, jangan dianggap error
                    continue
                line = raw.decode('utf-8', errors='ignore').strip()

                print("SERIAL >>>", line)

                # Firmware ESP32 mencampur baris debug teks biasa dengan baris
                # data JSON di stream Serial yang sama (lihat komentar di
                # RaspberryPiDataTransmitter.cpp) -> baris yang bukan JSON dilewati.
                if not line.startswith("{"):
                    continue

                try:
                    data = json.loads(line)
                except (json.JSONDecodeError, ValueError):
                    # Baris JSON terpotong (bisa terjadi di baud rate tinggi) -> lewati
                    continue

                self.last_raw_line = line

                # Format paket sesuai Transmitter_SendResult() di firmware:
                # {"rms_v":.., "rms_a":.., "cur":.., "temp":.., "vib_x":.., "vib_y":.., "vib_z":..,
                #  "rpm":.., "d2":.., "status":".."}
                self.current_v = float(data.get("rms_v", 0.0))
                self.current_a = float(data.get("rms_a", 0.0))
                self.current_cur = float(data.get("cur", 0.0))
                self.current_temp = float(data.get("temp", 0.0))
                # vib_x/y/z dipakai KHUSUS untuk tampilan grafik per-sumbu;
                # kalau firmware lama belum kirim field ini, default 0.0 saja
                # (bukan error) supaya tetap kompatibel dengan firmware lama.
                self.current_vx = float(data.get("vib_x", 0.0))
                self.current_vy = float(data.get("vib_y", 0.0))
                self.current_vz = float(data.get("vib_z", 0.0))
                self.current_rpm = float(data.get("rpm", 0.0))
                self.current_d2 = float(data.get("d2", 0.0))
                self.current_status_device = data.get("status", "")

                # Injeksi data masuk ke dalam antrean buffer internal grafik
                self.tick += 1
                self.time_buffer.append(self.tick)
                self.v_buffer.append(self.current_v)
                self.vx_buffer.append(self.current_vx)
                self.vy_buffer.append(self.current_vy)
                self.vz_buffer.append(self.current_vz)
                self.a_buffer.append(self.current_a)
                self.cur_buffer.append(self.current_cur)
                self.temp_buffer.append(self.current_temp)
                self.rpm_buffer.append(self.current_rpm)
                self.d2_buffer.append(self.current_d2)

                # Menyimpan data riil ke dalam berkas CSV jika perekaman aktif
                if self.recording and self.csv_writer:
                    # BUG LAMA: sebelumnya baris ini memakai variabel `elapsed`
                    # yang tidak pernah dihitung di fungsi ini -> setiap kali
                    # tombol recording ditekan, baris pertama yang masuk akan
                    # langsung lempar NameError, ketangkep except Exception di
                    # bawah, lalu thread serial dianggap "gagal tersambung"
                    # dan koneksi ke ESP32 di-reset berulang-ulang. Makanya
                    # recording kelihatan seperti tidak jalan / dashboard
                    # keputus-putus begitu recording dinyalakan.
                    elapsed = time.perf_counter() - self.record_start_time
                    self.csv_writer.writerow([
                        round(elapsed, 3),
                        self.machine_combo.currentText(),
                        self.current_v, self.current_a, self.current_cur, self.current_temp,
                        self.current_vx, self.current_vy, self.current_vz,
                        self.current_rpm, self.current_d2, self.current_status_device
                    ])
                    self.csv_file.flush()
            except Exception as e:
                print(f"[SERIAL] status: {e}")
                self.serial_connected = False
                self.ser = None
                self.current_v = self.current_a = self.current_cur = self.current_temp = None
                time.sleep(2)

    # ===================== EVALUASI DIAGNOSA MESIN (DIPAKAI TAB SUMMARY) =====================
    def _evaluate_diagnosis(self, v, temp, device_status=""):
        # Kalau firmware ESP32 sudah mengirim label status hasil klasifikasi
        # Mahalanobis + FFT (lebih akurat, lihat DiagnosisClassifier.cpp),
        # pakai itu langsung. Threshold v/temp di bawah cuma fallback kalau
        # firmware belum/tidak mengirim field "status".
        status_map = {
            "bahaya": ("STATUS: BAHAYA (CRITICAL)", COL_BAD,
                       "Terjadi anomali gesekan parah atau ketiadaan lubrikasi bearing! Segera matikan mesin produksi.",
                       "● DEVIASI BAHAYA"),
            "waspada": ("STATUS: WASPADA (WARNING)", COL_WARN,
                        "Indikasi awal ketidakseimbangan massa atau degradasi mekanis bearing terdeteksi.",
                        "● STATUS WASPADA"),
            "normal": ("STATUS: NORMAL", COL_OK,
                       "Seluruh parameter berjalan di bawah ambang batas deviasi krisis karsa cipta. Mesin aman digunakan.",
                       "● SYSTEM ONLINE"),
            # Dikirim firmware selama ~8 detik pertama setelah boot, selagi
            # buffer sensor (mic, FFT getaran, dll.) masih mengisi sample
            # pertamanya. Ini kondisi WAJAR, bukan error -> jangan ditampilkan
            # dengan warna merah/kuning yang bikin panik.
            "warming": ("STATUS: MENYIAPKAN SENSOR", "#888888",
                        "Perangkat baru menyala, sensor sedang mengambil sample pertama. Tunggu beberapa detik.",
                        "● WARMING UP"),
            # Status ini praktis tidak dikirim lagi oleh firmware versi terbaru
            # (main.ino sudah tidak punya fase kalibrasi), tapi tetap dijaga di
            # sini untuk kompatibilitas kalau firmware lama masih dipakai.
            "notcalibrated": ("STATUS: BELUM KALIBRASI", "#888888",
                               "Device belum menyelesaikan kalibrasi baseline awal. Biarkan mesin berjalan normal beberapa saat.",
                               "● KALIBRASI BASELINE"),
            "sensorfault": ("STATUS: SENSOR ERROR", COL_WARN,
                            "Data sensor basi/tidak lengkap terdeteksi firmware. Sudah lebih dari 8 detik sejak boot — cek sambungan sensor.",
                            "● SENSOR FAULT"),
        }
        key = (device_status or "").strip().lower()
        if key in status_map:
            title, color, desc, sys_txt = status_map[key]
            self.lbl_diag_status.setText(title)
            self.lbl_diag_status.setStyleSheet(f"font-size: 20px; font-weight: bold; color: {color};")
            self.box_diagnosis.setStyleSheet(f"background-color: transparent; border: 3px solid {color}; border-radius: 4px;")
            self.lbl_diag_desc.setText(desc)
            self.lbl_sys_status.setText(sys_txt)
            self.lbl_sys_status.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {color};")
            return

        # Evaluasi Kritis Kondisi Mesin (Diagnosa Otomatis di Tab Summary) - fallback
        if v > 0.25 or temp > 50.0:
            self.lbl_diag_status.setText("STATUS: BAHAYA (CRITICAL)")
            self.lbl_diag_status.setStyleSheet(f"font-size: 20px; font-weight: bold; color: {COL_BAD};")
            self.box_diagnosis.setStyleSheet(f"background-color: transparent; border: 3px solid {COL_BAD}; border-radius: 4px;")
            self.lbl_diag_desc.setText("Terjadi anomali gesekan parah atau ketiadaan lubrikasi bearing! Segera matikan mesin produksi.")
            status_txt, status_col = "● DEVIASI BAHAYA", COL_BAD
        elif v > 0.18 or temp > 42.0:
            self.lbl_diag_status.setText("STATUS: WASPADA (WARNING)")
            self.lbl_diag_status.setStyleSheet(f"font-size: 20px; font-weight: bold; color: {COL_WARN};")
            self.box_diagnosis.setStyleSheet(f"background-color: transparent; border: 3px solid {COL_WARN}; border-radius: 4px;")
            self.lbl_diag_desc.setText("Indikasi awal ketidakseimbangan massa atau degradasi mekanis bearing terdeteksi.")
            status_txt, status_col = "● STATUS WASPADA", COL_WARN
        else:
            self.lbl_diag_status.setText("STATUS: NORMAL")
            self.lbl_diag_status.setStyleSheet(f"font-size: 20px; font-weight: bold; color: {COL_OK};")
            self.box_diagnosis.setStyleSheet(f"background-color: transparent; border: 3px solid {COL_OK}; border-radius: 4px;")
            self.lbl_diag_desc.setText("Seluruh parameter berjalan di bawah ambang batas deviasi krisis karsa cipta. Mesin aman digunakan.")
            status_txt, status_col = "● SYSTEM ONLINE", COL_OK

        self.lbl_sys_status.setText(status_txt)
        self.lbl_sys_status.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {status_col};")

    def _on_machine_changed(self, text):
        # Dropdown-nya sekarang langsung nampilin nama mesin yang dipilih
        # (combo box-nya sendiri ada di header), jadi gak perlu lagi label
        # terpisah buat mengulang teks yang sama.
        pass

    def _reset_session(self):
        """Tombol RESET di header: nge-reset statistik & log kejadian anomali sesi
        berjalan saat ini. Tidak menghapus file rekaman CSV yang sudah tersimpan."""
        self.session_sample_count = 0
        self.session_rpm_sum = 0.0
        self.session_d2_max = 0.0
        self.session_worst_status = "Normal"
        self.session_waspada_count = 0
        self.session_bahaya_count = 0
        self.anomaly_events = []
        self.last_processed_tick = self.tick
        self.list_anomali.clear()
        self.list_anomali.addItem("Tidak ada kejadian anomali sepanjang sesi ini.")
        self._render_session_summary()

    def _show_debug_info(self):
        """Tombol DEBUG di header: nampilin info koneksi & baris data mentah
        terakhir yang diterima, buat troubleshooting cepat pas uji coba di lab."""
        port_info = self.ser.port if (self.ser is not None) else "(belum ada koneksi)"
        info = (
            f"Status koneksi   : {'TERSAMBUNG' if self.serial_connected else 'TIDAK TERSAMBUNG'}\n"
            f"Port serial      : {port_info}\n"
            f"Baud rate        : {BAUD_RATE}\n"
            f"Baris JSON akhir : {self.last_raw_line or '(belum ada data masuk)'}\n"
            f"Vib/Snd/Cur/Tmp  : {self.current_v}, {self.current_a}, {self.current_cur}, {self.current_temp}\n"
            f"RPM / D²         : {self.current_rpm}, {self.current_d2}\n"
            f"Status firmware  : {self.current_status_device or '-'}"
        )
        QMessageBox.information(self, "DEBUG - Info Koneksi & Data Terakhir", info)

    def _render_session_summary(self):
        self.lbl_session_summary.setText(
            f"Sesi: {self.session_sample_count} sample | "
            f"RPM rata-rata: {(self.session_rpm_sum / self.session_sample_count) if self.session_sample_count else 0.0:.1f} | "
            f"D² max: {self.session_d2_max:.2f} | "
            f"Kondisi terparah: {self.session_worst_status} | "
            f"Waspada: {self.session_waspada_count}x, Bahaya: {self.session_bahaya_count}x."
        )

    # ===================== LOGIKA EMBEDDED KOMPARASI & REFRESH GUI (MODE LIVE) =====================
    def _update_gui(self):
        # Update teks Jam digital di bagian baris atas
        now = datetime.now().strftime("%H:%M:%S")
        self.time_lbl.setText(now)

        # Indikator titik koneksi di header: hijau kalau tersambung, merah kalau tidak
        self.lbl_conn_dot.setStyleSheet(
            f"font-size: 12px; font-weight: bold; color: {COL_OK if self.serial_connected else COL_BAD};"
        )

        # Logika eksekusi jika data sensor valid masuk dari port serial
        if self.current_v is not None:
            # Refresh pergerakan 4 gelombang garis grafik secara simultan (Raw Reading)
            self.curve_v.setData(list(self.time_buffer), list(self.v_buffer))
            self.curve_a.setData(list(self.time_buffer), list(self.a_buffer))
            self.curve_cur.setData(list(self.time_buffer), list(self.cur_buffer))
            self.curve_temp.setData(list(self.time_buffer), list(self.temp_buffer))

            # Refresh grafik RPM & Mahalanobis D² (Processed)
            self.curve_rpm.setData(list(self.time_buffer), list(self.rpm_buffer))
            self.curve_d2.setData(list(self.time_buffer), list(self.d2_buffer))

            # Sinkronisasi teks angka di tab Raw Reading
            self.lbl_val_v.setText(f"Vib: {self.current_v:.2f} G")
            if self.current_vx is not None:
                self.lbl_val_vxyz.setText(
                    f"  (X: {self.current_vx:.2f} | Y: {self.current_vy:.2f} | Z: {self.current_vz:.2f} G)"
                )
            self.lbl_val_a.setText(f"Snd: {self.current_a:.1f} dB")
            self.lbl_val_cur.setText(f"Cur: {self.current_cur:.2f} A")
            self.lbl_val_temp.setText(f"Tmp: {self.current_temp:.1f} °C")
            if self.current_rpm is not None:
                self.lbl_val_rpm.setText(f"RPM: {self.current_rpm:.0f}")
            if self.current_d2 is not None:
                self.lbl_val_d2.setText(f"D² (Severity): {self.current_d2:.2f}")

            rpm_txt = f"{self.current_rpm:.0f}" if self.current_rpm is not None else "--"
            d2_txt = f"{self.current_d2:.2f}" if self.current_d2 is not None else "--"
            self.lbl_proc_snapshot.setText(
                f"Live | Vib: {self.current_v:.2f} G | Snd: {self.current_a:.1f} dB | "
                f"Cur: {self.current_cur:.2f} A | Tmp: {self.current_temp:.1f} °C | RPM: {rpm_txt} | D²: {d2_txt}"
            )

            # Evaluasi diagnosa otomatis. Kalau firmware sudah mengirim status_label
            # hasil klasifikasi Mahalanobis (Normal/Waspada/Bahaya), itu dipakai
            # langsung karena lebih akurat daripada re-threshold sederhana di Python.
            self._evaluate_diagnosis(self.current_v, self.current_temp, self.current_status_device)

            # Sinkronisasi panel besar di tab Summary
            self.lbl_diag_rpm.setText(f"RPM: {rpm_txt}")
            self.lbl_diag_d2.setText(f"D²: {d2_txt}")

            # ===== AKUMULASI STATISTIK SESI - hanya sekali per sample baru masuk,
            # bukan tiap 200ms timer, biar tidak dobel-hitung sample yang sama =====
            if self.tick != self.last_processed_tick:
                self.last_processed_tick = self.tick
                self.session_sample_count += 1
                if self.current_rpm is not None:
                    self.session_rpm_sum += self.current_rpm
                if self.current_d2 is not None:
                    self.session_d2_max = max(self.session_d2_max, self.current_d2)

                status_key = (self.current_status_device or "").strip().lower()
                if status_key == "waspada":
                    self.session_waspada_count += 1
                elif status_key == "bahaya":
                    self.session_bahaya_count += 1

                if status_key in STATUS_SEVERITY:
                    if STATUS_SEVERITY[status_key] > STATUS_SEVERITY.get(self.session_worst_status.lower(), 0):
                        self.session_worst_status = self.current_status_device.capitalize()

                if status_key in ("waspada", "bahaya"):
                    ts = datetime.now().strftime("%H:%M:%S")
                    event_txt = (
                        f"[{ts}] {self.current_status_device.upper()} — RPM {rpm_txt}, D² {d2_txt}, "
                        f"Vib {self.current_v:.2f}G, Tmp {self.current_temp:.1f}°C"
                    )
                    self.anomaly_events.append(event_txt)
                    if self.list_anomali.count() == 1 and self.list_anomali.item(0).text().startswith("Tidak ada"):
                        self.list_anomali.clear()
                    self.list_anomali.addItem(event_txt)
                    self.list_anomali.scrollToBottom()

                self._render_session_summary()
        elif not self.serial_connected:
            # Belum ada koneksi serial sama sekali -> tampilkan status jelas,
            # jangan biarkan panel diam di "STANDBY" tanpa penjelasan.
            self.lbl_sys_status.setText("● MENCARI PERANGKAT (SERIAL)...")
            self.lbl_sys_status.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_WARN};")
        else:
            # Port SUDAH konek, tapi belum ada baris JSON yang berhasil diparse
            # (misal ESP32 baru saja reboot dan masih mengirim baris teks debug
            # boot, bukan JSON) -> jangan tampilkan seolah masih gagal konek,
            # karena itu menyesatkan. Firmware sekarang TIDAK ADA lagi fase
            # kalibrasi 60 detik, jadi baris JSON pertama seharusnya muncul
            # dalam hitungan detik setelah boot.
            self.lbl_sys_status.setText("● TERSAMBUNG — MENUNGGU DATA JSON...")
            self.lbl_sys_status.setStyleSheet(f"font-size: 9px; font-weight: bold; color: {COL_ACCENT};")
            self.lbl_proc_snapshot.setText("Menunggu koneksi serial ke ESP32...")

    def _toggle_recording(self):
        if self.machine_combo.currentIndex() == 0:
            QMessageBox.warning(self, "Perhatian", "Silakan pilih Target Mesin terlebih dahulu!")
            return
            
        if not self.recording:
            filename = os.path.join(LOG_DIR, f"log_{datetime.now().strftime('%d%m%Y_%H%M%S')}.csv")
            try:
                self.csv_file = open(filename, 'w', newline='')
                self.csv_writer = csv.writer(self.csv_file)
                self.csv_writer.writerow(['timestamp', 'machine_type', 'rms_v', 'rms_a', 'current', 'temp', 'vib_x', 'vib_y', 'vib_z', 'rpm', 'mahalanobis_d2', 'status'])
                self.record_start_time = time.perf_counter()
                self.last_csv_time = 0.0
                self.recording = True
                self.btn_toggle_rec.setText("BERHENTI RECORDING")
                self.btn_toggle_rec.setStyleSheet(f"background-color: {COL_BAD}; color: #ffffff; font-weight: bold; font-size: 9px; height: 24px;")
                self.lbl_rec_status.setText(f"● MENULIS -> {os.path.basename(filename)}")
                self.lbl_rec_status.setStyleSheet(f"font-size: 8px; color: {COL_WARN};")
            except Exception as e:
                print(f"Gagal membuat file log: {e}")
        else:
            self.recording = False
            if self.csv_file:
                self.csv_file.close()
                self.csv_file = None
            self.btn_toggle_rec.setText("MULAI RECORDING")
            self.btn_toggle_rec.setStyleSheet("background-color: #cfcfcf; color: #000000; font-weight: bold; font-size: 9px; height: 24px;")
            self.lbl_rec_status.setText("● Berkas Tersimpan")
            self.lbl_rec_status.setStyleSheet(f"font-size: 8px; color: {COL_OK};")
            self._refresh_log_list()

    def _refresh_log_list(self):
        self.log_list.clear()
        if os.path.isdir(LOG_DIR):
            for file in sorted(os.listdir(LOG_DIR), reverse=True):
                if file.endswith(".csv"):
                    self.log_list.addItem(file)

    def _open_log_detail(self, item):
        """Dipanggil saat sebuah file rekaman di klik pada daftar -> menampilkan
        grafik hasil deteksi (Vibration/Sound/Current/Temp) beserta diagnosanya
        LANGSUNG di dashboard yang sama (halaman inline), bukan window/tab baru."""
        if item is None:
            return
        filename = item.text()
        filepath = os.path.join(LOG_DIR, filename)

        self.logdet_timer.stop()
        self.logdet_play_index = 0
        self.btn_logdet_play_pause.setText("▶ PLAY")
        self.lbl_logdet_title.setText(f"HASIL DETEKSI REKAMAN: {filename}")

        (self.logdet_times, self.logdet_v_vals, self.logdet_a_vals,
         self.logdet_cur_vals, self.logdet_t_vals) = self._logdet_load_csv(filepath)

        self.slider_logdet.setMaximum(max(0, len(self.logdet_times) - 1))
        self.slider_logdet.setValue(0)
        if self.logdet_times:
            # Tampilkan frame PALING AWAL saja (bukan seluruh data langsung),
            # jadi jelas kelihatan playback belum mulai/belum selesai — user
            # harus tekan PLAY dulu buat menjalankannya.
            self._logdet_render_frame(0)
        else:
            self.curve_logdet_v.clear()
            self.curve_logdet_a.clear()
            self.curve_logdet_cur.clear()
            self.curve_logdet_temp.clear()
            self.lbl_logdet_pos.setText("0 / 0")
        self._logdet_render_diagnosis_summary()

        self._change_page(4)

    def _open_log_detail_from_button(self):
        """Tombol cadangan 'Buka Panel Deteksi' - memakai file yang sedang terpilih di daftar."""
        current_item = self.log_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "Perhatian", "Pilih file rekaman (.csv) terlebih dahulu!")
            return
        self._open_log_detail(current_item)

    def _delete_selected_log(self):
        current_item = self.log_list.currentItem()
        if not current_item:
            return
            
        filename = current_item.text()
        filepath = os.path.join(LOG_DIR, filename)
        
        reply = QMessageBox.question(
            self, 'Konfirmasi Hapus', f"Apakah Anda yakin ingin menghapus berkas rekaman {filename}?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
                self._refresh_log_list()
                self.lbl_rec_status.setText("● BERKAS DIHAPUS")
                self.lbl_rec_status.setStyleSheet(f"font-size: 8px; color: {COL_BAD};")
            except Exception as e:
                print(f"Gagal menghapus file: {e}")

    def _export_selected_log_to_excel(self):
        """Mengubah berkas rekaman .csv yang dipilih di daftar Logs & Saves
        menjadi berkas Excel (.xlsx) yang rapi: header ditebalkan, kolom status
        diwarnai otomatis (hijau/jingga/merah), dan lebar kolom disesuaikan."""
        current_item = self.log_list.currentItem()
        if not current_item:
            QMessageBox.information(self, "Pilih Rekaman Dulu",
                                     "Klik salah satu berkas rekaman di daftar dulu sebelum export.")
            return

        if openpyxl is None:
            QMessageBox.critical(
                self, "Library Belum Terinstall",
                "Fitur export ke Excel butuh library 'openpyxl'.\n\n"
                "Install dulu lewat terminal/command prompt:\n\n"
                "    pip install openpyxl\n\n"
                "Setelah itu, jalankan ulang dashboard ini."
            )
            return

        filename = current_item.text()
        csv_path = os.path.join(LOG_DIR, filename)
        xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"

        try:
            with open(csv_path, 'r') as f:
                rows = list(csv.reader(f))

            if not rows:
                QMessageBox.warning(self, "Berkas Kosong", "Berkas rekaman ini tidak memiliki data.")
                return

            header, data_rows = rows[0], rows[1:]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Sensor ESP32"

            # ===== Tulis header (baris 1), ditebalkan + latar biru tua =====
            header_fill = PatternFill(start_color="1c1e22", end_color="1c1e22", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            ws.append(header)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Cari index kolom 'status' (kalau ada) untuk pewarnaan baris otomatis
            status_col_idx = None
            for i, col_name in enumerate(header):
                if col_name.strip().lower() == "status":
                    status_col_idx = i
                    break

            status_fill = {
                "normal":  PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid"),  # hijau muda
                "waspada": PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid"),   # jingga muda
                "bahaya":  PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid"),   # merah muda
            }

            # ===== Tulis data baris demi baris (coba konversi angka, biar Excel
            # membaca kolom sensor sebagai NUMBER, bukan teks) =====
            for row in data_rows:
                converted_row = []
                for value in row:
                    try:
                        converted_row.append(float(value))
                    except (ValueError, TypeError):
                        converted_row.append(value)  # kolom teks (mis. nama mesin, status) dibiarkan apa adanya
                ws.append(converted_row)

                if status_col_idx is not None:
                    status_val = str(row[status_col_idx]).strip().lower()
                    fill = status_fill.get(status_val)
                    if fill:
                        ws.cell(row=ws.max_row, column=status_col_idx + 1).fill = fill

            # ===== Lebar kolom otomatis menyesuaikan isi terpanjang =====
            for col_cells in ws.columns:
                max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
                ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2

            ws.freeze_panes = "A2"  # baris header tetap kelihatan pas di-scroll ke bawah

            wb.save(xlsx_path)

            QMessageBox.information(
                self, "Export Berhasil",
                f"Data berhasil diexport ke:\n{xlsx_path}\n\n"
                f"Total {len(data_rows)} baris data, siap dibuka di Microsoft Excel."
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Gagal", f"Terjadi kesalahan saat export ke Excel:\n{e}")


    def keyPressEvent(self, event):
        # Tekan tombol 'ESC' pada keyboard eksternal untuk menutup paksa aplikasi full-screen
        if event.key() == Qt.Key_Escape:
            if self.csv_file:
                self.csv_file.close()
            self.close()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    db = Dashboard()
    db.show()
    sys.exit(app.exec_())
